import pyshark
import openpyxl
import os
from shutil import move as smove
import glob
import dnp3_analiser as dnp3_an

filters = {         # Elimina o trafego entre a aplicação do coletor e o cc
    "COP-C-003": 'not ipv6.addr==fda0::4c:0:0:642d:7b1e',
    "COP-C-005": 'not ipv6.addr==fda0::12f:0:0:6416:e94e',
    "COP-C-006": 'not ipv6.addr==fda0::130:0:0:641a:752',       #TODO: Inserir novos coletores e criar pasta+planilha separada
    "REC-C-001": 'not ipv6.addr==fda0::bd:0:0:642c:f7f3',
    "REC-C-002": 'not ipv6.addr==fda0::ba:0:0:642f:4896',
    "REC-C-003": 'not ipv6.addr==fda0::bb:0:0:6440:8ae0',
    "REC-C-004": 'not ipv6.addr==fda0::bc:0:0:6440:8e57',
    "RBR-C-001": 'not ipv6.addr==fda0::fd:0:0:642b:d1b9',
    "RBR-C-002": 'not ipv6.addr==fda0::ff:0:0:6440:bfbf',
    "RBR-C-003": 'not ipv6.addr==fda0::fe:0:0:6440:d92b',
    "RBR-C-004": 'not ipv6.addr==fda0::135:0:0:6416:e95b',
    "RBR-C-005": 'not ipv6.addr==fda0::13f:0:0:6419:8adf',
    "RBR-C-007": 'not ipv6.addr==fda0::15e:0:0:642f:18fd',
    "RBR-C-008": 'not ipv6.addr==fda0::15d:0:0:6440:762',
    "BPD-C-001": 'not ipv6.addr==fda0::2f:0:0:642b:916c',
    "BPD-C-002": 'not ipv6.addr==fda0::af:0:0:6440:8e81',
    "BPD-C-003": 'not ipv6.addr==fda0::b0:0:0:6440:949f',
    "BPD-C-005": 'not ipv6.addr==fda0::111:0:0:6419:8b11',
    "BPD-C-006": 'not ipv6.addr==fda0::14d:0:0:6419:8b1e',
    "BPD-C-007": 'not ipv6.addr==fda0::14b:0:0:6419:8b2f',
    "BPD-C-008": 'not ipv6.addr==fda0::14c:0:0:6419:8b07'
}
file_format = ".pcap"
src = os.getcwd()
#src = r"C:\Users\klopffbe\Documents\Backup Bernardo\1000_LIGHT\1007_Analises LIGHT\Analises COP\20201020_Analise Wireshark (Romano)\Pcaps"
pcap_folder = src + "\\Pcaps"
path_planilha_geral = src + "\\Contagem_pacotes.xlsx"

rowNum = 0

if glob.glob(src+"\\*"+file_format):
    if not os.path.exists(pcap_folder):
        print(f"Creating {pcap_folder}")
        os.makedirs(pcap_folder)
    if not os.path.exists(path_planilha_geral):
        print(f"Creating {path_planilha_geral}...")
        workBook = openpyxl.Workbook()
        dataSheet = workBook.active
        dataSheet.title = "Contagem"
        dataSheet.cell(row=1, column=1).value = "Data"
        dataSheet.cell(row=1, column=2).value = "Coletor"
        dataSheet.cell(row=1, column=3).value = "DNP3"
        dataSheet.cell(row=1, column=4).value = "TCP"
        dataSheet.cell(row=1, column=5).value = "UDP"
        dataSheet.cell(row=1, column=6).value = "ICMPv6"
        dataSheet.cell(row=1, column=7).value = "UAUDP"
        dataSheet.cell(row=1, column=8).value = "PANA"
        dataSheet.cell(row=1, column=9).value = "SNMP"
        dataSheet.cell(row=1, column=10).value = "Total"
        rowNum = 1
    else:
        print(f"Opening {path_planilha_geral}...")
        workBook = openpyxl.load_workbook(path_planilha_geral)
        if "Contagem" in workBook.sheetnames:
            dataSheet = workBook["Contagem"]
        else:
            dataSheet = workBook.create_sheet("Contagem", 0)
            dataSheet.cell(row=1, column=1).value = "Data"
            dataSheet.cell(row=1, column=2).value = "Coletor"
            dataSheet.cell(row=1, column=3).value = "DNP3"
            dataSheet.cell(row=1, column=4).value = "TCP"
            dataSheet.cell(row=1, column=5).value = "UDP"
            dataSheet.cell(row=1, column=6).value = "ICMPv6"
            dataSheet.cell(row=1, column=7).value = "UAUDP"
            dataSheet.cell(row=1, column=8).value = "PANA"
            dataSheet.cell(row=1, column=9).value = "SNMP"
            dataSheet.cell(row=1, column=10).value = "Total"
        rowNum = dataSheet.max_row                  # Gives me the least used line
    for file in glob.glob(src+"\\*"+file_format):   # fixme: with os.path.join()
        file_name = os.path.basename(file)
        file_colector = file_name[5:14]                           # re.findall("[A-Z]{3}-[A-Z]-[0-9]{3}", file_name)
        file_date = f"{file_name[21:23]}/{file_name[19:21]}/{file_name[15:19]}"
        if not file_colector in filters:
            print(f"ERROR: There is no stored filter for {file_colector}!")
            pass
        try:
            pcap_file = pyshark.FileCapture(file, only_summaries=True, display_filter=filters[file_colector])
            print(f"Parseando {file_name}...")
            TCP = 0
            DNP3 = 0
            PANA = 0
            UDP = 0
            ICMPv6 = 0
            SNMP = 0
            UAUDP = 0
            for pkt in pcap_file:
                if pkt.protocol == 'TCP':
                    # print(pkt.protocol)
                    TCP = TCP + 1
                elif pkt.protocol == 'ICMPv6':
                    # print(pkt.protocol)
                    ICMPv6 = ICMPv6 + 1
                elif pkt.protocol == 'DNP 3.0':
                    # print(pkt.protocol)
                    DNP3 = DNP3 + 1
                elif pkt.protocol == 'PANA':
                    # print(pkt.protocol)
                    PANA = PANA + 1
                elif pkt.protocol == 'UDP':
                    # print(pkt.protocol)
                    UDP = UDP + 1
                elif pkt.protocol == 'SNMP':
                    # print(pkt.protocol)
                    SNMP = SNMP + 1
                elif pkt.protocol == 'UAUDP':
                    # print(pkt.protocol)
                    UAUDP = UAUDP + 1
            pcap_file.close()
            rowNum += 1
            dataSheet.cell(row=rowNum, column=1).value = file_date
            print("Contagem no coletor", file_colector)
            dataSheet.cell(row=rowNum, column=2).value = file_colector
            print("Quantidade de pacotes DNP3:", DNP3)
            dataSheet.cell(row=rowNum, column=3).value = DNP3
            print("Quantidade de pacotes TCP:", TCP)
            dataSheet.cell(row=rowNum, column=4).value = TCP
            print("Quantidade de pacotes UDP:", UDP)
            dataSheet.cell(row=rowNum, column=5).value = UDP
            print("Quantidade de pacotes ICMPv6:", ICMPv6)
            dataSheet.cell(row=rowNum, column=6).value = ICMPv6
            print("Quantidade de pacotes UAUDP:", UAUDP)
            dataSheet.cell(row=rowNum, column=7).value = UAUDP
            print("Quantidade de pacotes PANA:", PANA)
            dataSheet.cell(row=rowNum, column=8).value = PANA
            print("Quantidade de pacotes SNMP:", SNMP)
            dataSheet.cell(row=rowNum, column=9).value = SNMP
            dataSheet.cell(row=rowNum, column=10).value = f"=SUM(C{rowNum}: I{rowNum})"
        except:
            print(f"Erro ao Parsear {file_name}! Arquivo PCAP pode estar corrompido...")
            # pcap_file.close()
            pass
        workBook.save(path_planilha_geral)
        dnp3_an.dnp3_packet(file, file_colector)
        date_folder = pcap_folder + f"\\{file_name[19:21]}_{file_name[21:23]}"
        if not os.path.exists(date_folder):
            os.makedirs(date_folder)
        smove(file, date_folder + "\\" + file_name)
    workBook.close()
    try:
        dnp3_an.append_df_to_excel(filename=path_planilha_geral, df=dnp3_an.create_dataframe(dnp3_an.list_export),
                                   sheet_name='DNP3', index=False)
    except:
        print(f"ERROR: Ao exportar as leituras DNP3 para .xlsx")
        pass
    print("Done!")

else:
    print("No .PCAP files found!\nExiting...")
