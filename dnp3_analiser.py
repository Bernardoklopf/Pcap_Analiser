import pyshark                                                  # https://github.com/KimiNewt/pyshark/
import os
import pandas as pd
import glob
import datetime
run_time = datetime.datetime.now()
# pcap_file = pyshark.FileCapture(file, display_filter="(dnp3) && (ipv6.addr==fda0::ff:1c:6400:8077:2213)", keep_packets=True)
#                                                                 # param keep_packets: Whether to keep packets after reading
#                                                                 # them via next(). Used to conserve memory when reading large caps.
# # Todo: Usar Only summaries para extrair tipo de resposta.
# # pcap_file = pyshark.FileCapture(file, only_summaries=True, display_filter="(dnp3) && (dnp3.al.uns==1)")
#
# packet = pcap_file[1]                                           # https://www.wireshark.org/docs/dfref/d/dnp3.html


# pcap_file2 = pyshark.FileCapture(file2, display_filter="(dnp3) && (ipv6.addr==fda0::ff:1c:6400:8077:2213)")
# dnp3_func_2 = []
# ... for packet in pcap_file2:
# ...     if not packet.dnp3.al_func in dnp3_func2:
# ...         print(packet.dnp3.al_func)
# ...         dnp3_func2.append(packet.dnp3.al_func)

# def return_packet(packet_no, pcap_file):
#      for packet in pcap_file:
#          if packet.number == str(packet_no):
#              return packet
#          else: pass
#      print("packet not found")


# def return_point_numbers(packet):
#     point_number_list = []
#     for index in packet.dnp3.al_index.all_fields:  # retorna todos os point numbers A+D
#         if index.hex_value not in point_number_list:
#             point_number_list.append(index.hex_value)
#     return point_number_list


# class Endpoint:         #Todo: Check if it's better to use a dictionary or a Class
#     def __init__(self, SerialNo, Colector=None, Dnp3_total=None, Uns_Resp=None, Read_Analog_Input=None, Read_Class=None, Responses=None,
#                  Unsolicited_Response=None, TCP_Retransmissions=None, Reset_of_Remote_Link=None):
#         self.SerialNo = SerialNo
#         self.Colector = Colector
#         self.Dnpr_total = Dnp3_total
#         self.Uns_Resp = Uns_Resp
#         self.Read_Analog_Input = Read_Analog_Input
#         self.Read_Class = Read_Class
#         self.Responses = Responses
#         self.Unsolicited_Response = Unsolicited_Response
#         self.TCP_Retransmissions = TCP_Retransmissions
#         self.Reset_of_Remote_Link = Reset_of_Remote_Link
#
#     def add_Unsolicited_Response(self):
#         self.Unsolicited_Response +=1
#
# dic={}
# dic[str(88779966)]= Endpoint(88779966, Unsolicited_Response=0)  # Todo: Utilizar Dicionario de objetos.
#
# def create_endpoint(_SerialNo):
#     return Endpoint(SerialNo=_SerialNo)

# def dnp3_analise(packet):

# print(f"Opening {path_planilha_geral}...")
# workBook = openpyxl.load_workbook(path_planilha_geral)
# if "Point Numbers" in workBook.sheetnames:
#     dataSheet = workBook["Point Numbers"]
# else:
#     dataSheet = workBook.create_sheet("Point Numbers")
#     dataSheet.cell(row=1, column=1).value = "Coletor"
#     dataSheet.cell(row=1, column=2).value = "Serial"
#     dataSheet.cell(row=1, column=3).value = "DNP3 Src"
#     dataSheet.cell(row=1, column=4).value = "DNP3 Dst"
#     dataSheet.cell(row=1, column=5).value = "Quantidade de PNs"
#     dataSheet.cell(row=1, column=6).value = "Point Numbers"
# rowNum = dataSheet.max_row

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,    # https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    from openpyxl import load_workbook
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
            to_excel_kwargs["header"] = None
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()





# src = r"C:\Users\klopfflb\OneDrive - Landis+Gyr\1000_LIGHT\1007_Analises LIGHT\Analises BPD\2020_08_17-Analise RL86813 e\RL86813"
# path_planilha_geral = r"C:\Users\klopfflb\OneDrive - Landis+Gyr\1000_LIGHT\1007_Analises LIGHT\Analises BPD\2020_08_17-Analise RL86813 e\RL86813"
# file_format = ".pcap"


display_filter = "(dnp3)"
cabecalho = ["Arquivo", "Coletor", "Pkt N.", "Pkt Date", "Pkt Time", "Serial N.", "Dnp3 Src.", "Dnp3 Dst",
                 "Func. Code (Tipo de Pacote)", "Point N. Function (A/D)", "Quantity", "Point Numbers"]
list_export = []


def create_dataframe(lista):
    print("Creating data Frame")
    data_frame = pd.DataFrame(lista)
    print("Insterting Cabeçalho.")
    data_frame.columns = cabecalho
    return data_frame


def create_file(f_name, header, lista_dic):
    print("Creating data Frame")
    data_frame = pd.DataFrame(lista_dic)
    print("Insterting Cabeçalho.")
    data_frame.columns = cabecalho
    data_frame.to_excel(f"{f_name}.xlsx", index=False)
    data_frame.to_csv(f"{f_name}.csv", index=False)


# def dnp3_analisis(packet, workbook):
#     if "DNP3" in workBook.sheetnames:
#         dataSheet = workBook["DNP3"]
#     else:
#         dataSheet = workBook.create_sheet("DNP3", 1)
#         for value, info in enumerate(cabecalho):
#             dataSheet.cell(row=1,column=value+1).value = info
#     rowNum = dataSheet.max_row  # Gives me the least used line


def dnp3_packet(file, coletor):
    Counter = 0
    file_name = os.path.basename(file)
    print(f"Parseando pacotes DNP3 em {file_name}...")
    dic_export = dict(Arquivo=None, Coletor=None, Pkt_No=None, Date=None, Time=None, SerialNo=None, Dnp3_src=None,
                      Dnp3_dst=None, Function_code=None, Obj_ad_function=None, Obj_Qnt=None, Pn_string="")
    pcap_file = pyshark.FileCapture(file, display_filter=display_filter, keep_packets=False)
    for packet in pcap_file:
        # print(f"Lendo Pacote {packet.number}")
        try:
            dic_export["Arquivo"] = file_name
            dic_export["Coletor"] = coletor
            dic_export["Pkt_No"] = packet.number
            dic_export["Date"] = packet.sniff_time.strftime("%m/%d/%Y")  # general date
            dic_export["Time"] = packet.sniff_time.strftime("%H:%M:%S")  # general time
            dic_export["SerialNo"] = packet.ipv6.src[-9:].replace(":", "")  # Radio
            dic_export["Dnp3_src"] = packet.dnp3.src  # endereço DNP3
            dic_export["Dnp3_dst"] = packet.dnp3.dst  # endereço saida DNP3
            dic_export["Obj_Qnt"] = packet.dnp3.al_range_quantity  # quantidade de alterações
            dic_export["Function_code"] = hex(int(packet.dnp3.al_func))  # Informação da função do Pacote (vide tabela DNP3)
            dic_export["Obj_ad_function"] = hex(
                int(packet.dnp3.al_obj))  # retorna uma Str do Hex do Obj. Func. Para saber se é A/D
            try:
                # print(f"Extraindo Point Numbers")
                # point_number_list = return_point_numbers(packet)   # retorna todos os point numbers A+D
                # point_number_list = []                             # Todo: Activate Analog and Digital read values .dnp3.al_ana_int.all_fields
                point_number_dic = {}
                for point_no in packet.dnp3.al_index.all_fields:  # retorna todos os point numbers A+D
                    if str(point_no.hex_value) not in point_number_dic:
                        point_number_dic[str(point_no.hex_value)] = 1
                        # point_number_list.append(index.hex_value)
                    else:
                        point_number_dic[str(point_no.hex_value)] += 1
                dic_export["Pn_string"] = ""
                for pn in point_number_dic:
                    dic_export["Pn_string"] += f"{pn}x{point_number_dic[pn]}, "
            except:
                # print("No point numbers availabe.")
                dic_export["Pn_string"] = ""
            Counter += 1
            dic_temp = dic_export.copy()
            list_export.append(dic_temp)
        except:
            # print(f"Skipping packet {packet.number}")
            pass
    print(f"{Counter} Packets Extracted.")
    pcap_file.close()



def dnp3_analisys_alone(src, file_format):
    dic_export = dict(Arquivo=None, Pkt_No=None, Date=None, Time=None, SerialNo=None, Dnp3_src=None, Dnp3_dst=None,
                      Function_code=None, Obj_ad_function=None, Obj_Qnt=None, Pn_string="")
    file_name = ""
    for file in glob.glob(src + "\\*" + file_format):
        # list_export.clear()
        file_name = os.path.basename(file)
        # file_colector = file_name[5:14]
        # file_date = f"{file_name[21:23]}/{file_name[19:21]}/{file_name[15:19]}"
        print(f"Abrindo {file_name}")
        pcap_file = pyshark.FileCapture(file, display_filter=display_filter, keep_packets=True)
        #try:
        print(f"Parseando {file_name}...")
        for packet in pcap_file:
            print(f"Lendo Pacote {packet.number}")
            try:
                dic_export["Arquivo"] = file_name
                dic_export["Pkt_No"] = packet.number
                dic_export["Date"] = packet.sniff_time.strftime("%m/%d/%Y")          # general date
                dic_export["Time"] = packet.sniff_time.strftime("%H:%M:%S")          # general time
                dic_export["SerialNo"] = packet.ipv6.src[-9:].replace(":", "")       # Radio
                dic_export["Dnp3_src"] = packet.dnp3.src                             # endereço DNP3
                dic_export["Dnp3_dst"] = packet.dnp3.dst                             # endereço saida DNP3
                dic_export["Obj_Qnt"] = packet.dnp3.al_range_quantity                # quantidade de alterações
                dic_export["Function_code"] = hex(int(packet.dnp3.al_func))          # Informação da função do Pacote (vide tabela DNP3)
                dic_export["Obj_ad_function"] = hex(int(packet.dnp3.al_obj))         # retorna uma Str do Hex do Obj. Func. Para saber se é A/D
                try:
                    print(f"Extraindo Point Numbers")
                    # point_number_list = return_point_numbers(packet)                   # retorna todos os point numbers A+D
                    # point_number_list = []                                             # Todo: Activate Analog and Digital read values .dnp3.al_ana_int.all_fields
                    point_number_dic = {}
                    for point_no in packet.dnp3.al_index.all_fields:                     # retorna todos os point numbers A+D
                        if str(point_no.hex_value) not in point_number_dic:
                            point_number_dic[str(point_no.hex_value)] = 1
                            # point_number_list.append(index.hex_value)
                        else:
                            point_number_dic[str(point_no.hex_value)] += 1
                    dic_export["Pn_string"] = ""
                    for pn in point_number_dic:
                        dic_export["Pn_string"] += f"{pn}x{point_number_dic[pn]}, "
                except:
                    print("No point numbers availabe.")
                    dic_export["Pn_string"] = ""
                dic_temp = dic_export.copy()
                list_export.append(dic_temp)
            except:
                print(f"Skipping packet {packet.number}")
                pass
    # create_file(f"Dnp3_Analisys_{run_time.strftime('%Y_%m_%d_%Hh%Mm')}",cabecalho,list_export)
